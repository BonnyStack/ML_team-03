from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "archive"

INPUT_FILE = DATA_DIR / "olist_merged_dataset (2).xlsx"
OUTPUT_DIR = DATA_DIR / "split_dataset"
TRAIN_RATIO = 0.8
VALIDATION_RATIO = 0.1
TEST_RATIO = 0.1


def log(message: str) -> None:
    print(message, flush=True)


def count_data_rows(sheet) -> int:
    total_rows = 0
    for row_index, _ in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_index == 1:
            continue
        total_rows += 1
    return total_rows


def build_split_index_sets(row_count: int) -> tuple[set[int], set[int], set[int]]:
    train_count = int(row_count * TRAIN_RATIO)
    validation_count = int(row_count * VALIDATION_RATIO)
    test_count = row_count - train_count - validation_count

    train_rows = set(range(1, train_count + 1))
    validation_rows = set(range(train_count + 1, train_count + validation_count + 1))
    test_rows = set(
        range(
            train_count + validation_count + 1,
            train_count + validation_count + test_count + 1,
        )
    )
    return train_rows, validation_rows, test_rows


def create_output_workbooks() -> tuple[Workbook, Workbook, Workbook]:
    train_wb = Workbook(write_only=True)
    validation_wb = Workbook(write_only=True)
    test_wb = Workbook(write_only=True)
    return train_wb, validation_wb, test_wb


def remove_default_sheet_if_needed(workbook: Workbook) -> None:
    if workbook.worksheets and workbook.worksheets[0].title == "Sheet":
        workbook.remove(workbook.worksheets[0])


def split_workbook() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"File not found: {INPUT_FILE}")

    if abs((TRAIN_RATIO + VALIDATION_RATIO + TEST_RATIO) - 1.0) > 1e-9:
        raise ValueError("Split ratios must add up to 1.0")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    log(f"Opening source workbook: {INPUT_FILE}")
    source_wb = load_workbook(INPUT_FILE, read_only=True, data_only=False)

    train_wb, validation_wb, test_wb = create_output_workbooks()
    remove_default_sheet_if_needed(train_wb)
    remove_default_sheet_if_needed(validation_wb)
    remove_default_sheet_if_needed(test_wb)

    split_plan: dict[str, tuple[set[int], set[int], set[int]]] = {}

    for sheet in source_wb.worksheets:
        data_row_count = count_data_rows(sheet)
        split_plan[sheet.title] = build_split_index_sets(data_row_count)

        train_rows, validation_rows, test_rows = split_plan[sheet.title]
        log(
            f"Planned split for '{sheet.title}': "
            f"train {len(train_rows):,}, "
            f"validation {len(validation_rows):,}, "
            f"test {len(test_rows):,}"
        )

    source_wb.close()

    log("Re-opening workbook to write split files")
    source_wb = load_workbook(INPUT_FILE, read_only=True, data_only=False)

    total_train = 0
    total_validation = 0
    total_test = 0

    for sheet in source_wb.worksheets:
        train_sheet = train_wb.create_sheet(title=sheet.title)
        validation_sheet = validation_wb.create_sheet(title=sheet.title)
        test_sheet = test_wb.create_sheet(title=sheet.title)

        train_rows, validation_rows, test_rows = split_plan[sheet.title]
        data_row_index = 0

        for row_index, row_values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                train_sheet.append(row_values)
                validation_sheet.append(row_values)
                test_sheet.append(row_values)
                continue

            data_row_index += 1

            if data_row_index in train_rows:
                train_sheet.append(row_values)
                total_train += 1
            elif data_row_index in validation_rows:
                validation_sheet.append(row_values)
                total_validation += 1
            else:
                test_sheet.append(row_values)
                total_test += 1

    source_wb.close()

    train_file = OUTPUT_DIR / "olist_train_80.xlsx"
    validation_file = OUTPUT_DIR / "olist_validation_10.xlsx"
    test_file = OUTPUT_DIR / "olist_test_10.xlsx"

    log(f"Saving training file: {train_file}")
    train_wb.save(train_file)
    log(f"Saving validation file: {validation_file}")
    validation_wb.save(validation_file)
    log(f"Saving testing file: {test_file}")
    test_wb.save(test_file)

    log(
        f"Done. Training rows: {total_train:,}, "
        f"validation rows: {total_validation:,}, "
        f"testing rows: {total_test:,}"
    )
    log(f"Split files saved in: {OUTPUT_DIR}")


if __name__ == "__main__":
    split_workbook()
