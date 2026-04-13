from pathlib import Path
from openpyxl import Workbook, load_workbook


PROGRESS_EVERY = 5_000


def create_copy_name(file_path: Path) -> Path:
    """
    Create a new filename in the same folder with ' (2)' before the extension.
    Example:
    olist_merged_dataset.xlsx -> olist_merged_dataset (2).xlsx
    """
    return file_path.with_name(f"{file_path.stem} (2){file_path.suffix}")


def row_is_complete(values) -> bool:
    """
    Return True only if every cell in the row has a value.
    Empty string and whitespace-only strings are treated as missing.
    """
    for value in values:
        if value is None:
            return False
        if isinstance(value, str) and value.strip() == "":
            return False
    return True


def log(message: str) -> None:
    print(message, flush=True)


def copy_complete_rows(source_sheet, target_sheet) -> tuple[int, int]:
    total_rows = 0
    kept_rows = 0

    for row_values in source_sheet.iter_rows(values_only=True):
        total_rows += 1

        if row_is_complete(row_values):
            target_sheet.append(row_values)
            kept_rows += 1

        if total_rows % PROGRESS_EVERY == 0:
            log(
                f"  Processed {total_rows:,} rows "
                f"(kept {kept_rows:,}, removed {total_rows - kept_rows:,})"
            )

    return total_rows, kept_rows


def clean_xlsx_remove_incomplete_rows(input_file: str):
    input_path = Path(input_file)

    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")

    output_path = create_copy_name(input_path)

    log(f"Opening workbook: {input_path}")
    source_wb = load_workbook(input_path, read_only=True, data_only=False)
    target_wb = Workbook(write_only=True)

    total_rows_all_sheets = 0
    kept_rows_all_sheets = 0

    for sheet_index, source_sheet in enumerate(source_wb.worksheets, start=1):
        log(
            f"Processing sheet {sheet_index}/{len(source_wb.worksheets)}: "
            f"{source_sheet.title}"
        )

        target_sheet = target_wb.create_sheet(title=source_sheet.title)
        total_rows, kept_rows = copy_complete_rows(source_sheet, target_sheet)

        removed_rows = total_rows - kept_rows
        total_rows_all_sheets += total_rows
        kept_rows_all_sheets += kept_rows

        log(
            f"Finished sheet '{source_sheet.title}': "
            f"total {total_rows:,}, kept {kept_rows:,}, removed {removed_rows:,}"
        )

    log(f"Saving cleaned workbook to: {output_path}")
    target_wb.save(output_path)
    source_wb.close()

    removed_all_sheets = total_rows_all_sheets - kept_rows_all_sheets
    log(
        f"Done. Total rows {total_rows_all_sheets:,}, "
        f"kept {kept_rows_all_sheets:,}, removed {removed_all_sheets:,}"
    )
    log(f"Cleaned file saved here: {output_path}")


if __name__ == "__main__":
    file_location = r"C:\Users\alber\Downloads\archive\olist_merged_dataset.xlsx"
    clean_xlsx_remove_incomplete_rows(file_location)
