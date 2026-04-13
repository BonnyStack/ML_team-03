from pathlib import Path

import matplotlib.pyplot as plt
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "archive"

INPUT_FILE = DATA_DIR / "olist_merged_dataset (2).xlsx"
OUTPUT_FILE = DATA_DIR / "olist_is_delayed_pie_chart.png"
TARGET_COLUMN = "is_delayed"


def log(message: str) -> None:
    print(message, flush=True)


def normalize_delay_value(value) -> int | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        if value in (0, 1):
            return int(value)
        return None

    text = str(value).strip().lower()
    if text in {"0", "0.0", "false", "no"}:
        return 0
    if text in {"1", "1.0", "true", "yes"}:
        return 1
    return None


def find_column_index(header_row: tuple) -> int:
    for index, value in enumerate(header_row):
        if str(value).strip() == TARGET_COLUMN:
            return index
    raise ValueError(f"Column '{TARGET_COLUMN}' not found in the workbook.")


def count_delayed_orders() -> tuple[int, int]:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"File not found: {INPUT_FILE}")

    workbook = load_workbook(INPUT_FILE, read_only=True, data_only=False)

    delayed_count = 0
    not_delayed_count = 0

    for sheet in workbook.worksheets:
        log(f"Reading sheet: {sheet.title}")
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)

        if header_row is None:
            continue

        column_index = find_column_index(header_row)

        for row in rows:
            delay_value = normalize_delay_value(row[column_index])

            if delay_value == 1:
                delayed_count += 1
            elif delay_value == 0:
                not_delayed_count += 1

    workbook.close()
    return delayed_count, not_delayed_count


def save_pie_chart(delayed_count: int, not_delayed_count: int) -> None:
    labels = ["Delayed orders", "Not delayed orders"]
    sizes = [delayed_count, not_delayed_count]
    colors = ["#FFF4C3", "#4643FF"]

    plt.figure(figsize=(7, 7))
    plt.pie(
        sizes,
        labels=labels,
        colors=colors,
        autopct="%1.1f%%",
        startangle=90,
        counterclock=False,
    )
    plt.title("Delayed vs Not Delayed Orders")
    plt.axis("equal")
    plt.tight_layout()
    plt.savefig(OUTPUT_FILE, dpi=200, bbox_inches="tight")
    plt.close()

    log(f"Pie chart saved to: {OUTPUT_FILE}")


def main() -> None:
    delayed_count, not_delayed_count = count_delayed_orders()
    total = delayed_count + not_delayed_count

    if total == 0:
        raise ValueError("No valid is_delayed values (0 or 1) were found in the workbook.")

    log(
        f"Delayed orders: {delayed_count:,} | "
        f"Not delayed orders: {not_delayed_count:,}"
    )
    save_pie_chart(delayed_count, not_delayed_count)


if __name__ == "__main__":
    main()
