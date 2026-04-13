from argparse import ArgumentParser, Namespace
from pathlib import Path

import matplotlib.pyplot as plt
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "archive"

DEFAULT_ORIGINAL_FILE = DATA_DIR / "olist_merged_dataset.xlsx"
DEFAULT_CLEANED_FILE = DATA_DIR / "olist_merged_dataset (2).xlsx"
DEFAULT_OUTPUT_FILE = DATA_DIR / "olist_cleanup_pie_chart.png"

def log(message: str) -> None:
    print(message, flush=True)


def count_rows(workbook_path: Path) -> int:
    if not workbook_path.exists():
        raise FileNotFoundError(f"File not found: {workbook_path}")

    log(f"Opening workbook: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)

    total_rows = 0
    for sheet in workbook.worksheets:
        sheet_rows = sum(1 for _ in sheet.iter_rows(values_only=True))
        total_rows += sheet_rows
        log(f"Sheet '{sheet.title}': {sheet_rows:,} rows")

    workbook.close()
    return total_rows


def save_pie_chart(kept_rows: int, removed_rows: int, output_file: Path) -> None:
    labels = ["Kept rows", "Removed rows"]
    sizes = [kept_rows, removed_rows]
    colors = ["#2A9D8F", "#E76F51"]

    plt.figure(figsize=(7, 7))
    plt.pie(
        sizes,
        labels=labels,
        colors=colors,
        autopct="%1.1f%%",
        startangle=90,
        counterclock=False,
    )
    plt.title("Rows Kept vs Removed")
    plt.axis("equal")
    plt.tight_layout()
    plt.savefig(output_file, dpi=200, bbox_inches="tight")
    plt.close()

    log(f"Pie chart saved to: {output_file}")


def parse_args() -> Namespace:
    parser = ArgumentParser(description="Create a pie chart for kept vs removed rows.")
    parser.add_argument(
        "--original",
        type=Path,
        default=DEFAULT_ORIGINAL_FILE,
        help="Path to the original workbook.",
    )
    parser.add_argument(
        "--cleaned",
        type=Path,
        default=DEFAULT_CLEANED_FILE,
        help="Path to the cleaned workbook.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help="Path to save the pie chart image.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    original_rows = count_rows(args.original)
    kept_rows = count_rows(args.cleaned)
    removed_rows = original_rows - kept_rows

    if removed_rows < 0:
        raise ValueError(
            "The cleaned workbook has more rows than the original workbook. "
            "Check that the selected files are correct."
        )

    log(
        f"Total rows: {original_rows:,} | "
        f"Kept rows: {kept_rows:,} | Removed rows: {removed_rows:,}"
    )
    save_pie_chart(kept_rows, removed_rows, args.output)


if __name__ == "__main__":
    main()
